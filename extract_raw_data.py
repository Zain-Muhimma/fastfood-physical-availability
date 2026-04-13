"""
Extract additional data from raw respondent-level Excel
'Burgerizzr BHT_ WAVE 3 complete labels.xlsx'

Produces supplementary JSON that gets merged into pa_data.json:
- All 15 Q24 attributes (currently only 7 extracted from % sheet)
- Per-brand Q25 reasons (cross-tabulated with Q15b main brand)
- Q3 channel frequency
- Q15b main brand market share
"""
import json, sys, os
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

RAW_FILE = 'Burgerizzr BHT_ WAVE 3 complete labels.xlsx'
OUTPUT = 'public/data/pa_data.json'

TARGET_BRANDS = [
    "Kudu", "Burgerizzr", "Al Baik", "McDonalds", "KFC",
    "Sign Burger", "Burger King", "Bayt Al-Shawarma", "Hardees", "Herfy"
]

# Brand name normalization
BRAND_NORM = {
    "McDonald's": "McDonalds", "Hardee's": "Hardees", "I'm Hungry": "Im Hungry",
    "Chef's": "Chefs", "Raising Cane's": "Raising Canes", "Popeye's": "Popeyes",
    "Domino's Pizza": "Dominos Pizza", "Im Hungry": "Im Hungry",
}

def norm(name):
    if not name: return name
    name = str(name).strip().replace('\u2019', "'")
    return BRAND_NORM.get(name, name)

def is_target(name):
    return norm(name) in TARGET_BRANDS

# Q24 attribute column ranges (1-indexed from the header exploration)
# Each Q24 attribute spans 36 columns (35 brands + None)
Q24_ATTRS = {
    'Taste': 1296, 'Quality': 1332, 'Price': 1368, 'Promos': 1404,
    'Variety': 1440, 'Health': 1476, 'Location': 1512, 'Ambiance': 1548,
    'Clientele': 1584, 'Cleanliness': 1620, 'Service': 1656, 'Speed': 1692,
    'Reputation': 1728, 'Popularity': 1764, 'Innovation': 1800,
}

# Q25 reason columns (1-indexed)
Q25_COLS = {
    'Effortless choice': 1836, 'Personal connection': 1837,
    'Social influence': 1838, 'Unique menu appeal': 1839,
    'Positive physical feeling': 1840, 'Emotional comfort': 1841,
    'Location convenience': 1842, 'Routine': 1843,
    'Cultural fit': 1844, 'Consistent promotions': 1845,
    'Customizable experience': 1846, 'Lack of alternatives': 1847,
    'Reliable ordering': 1848, 'Branch loyalty': 1849,
    'Feeling special': 1850,
}

# Q3 channel columns
Q3_COLS = {
    'Dine-in': 84, 'Takeaway': 85, 'Drive-through': 86,
    'Restaurant app': 87, 'Aggregator app': 88,
}

# Q15b main brand column
Q15B_COL = 332


def main():
    print(f"Loading {RAW_FILE}...")
    wb = openpyxl.load_workbook(RAW_FILE, read_only=True, data_only=True)
    ws = wb['complete labels']

    # Read header row to get brand names for Q24 columns
    header = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    total_cols = len(header)
    print(f"Total columns: {total_cols}")

    # Build Q24 brand name mapping for each attribute
    q24_brand_maps = {}
    for attr, start_col in Q24_ATTRS.items():
        brands_in_attr = []
        for c in range(start_col - 1, min(start_col + 35, total_cols)):
            h = str(header[c]) if header[c] else ''
            # Extract brand name from header like "McDonalds - For each..."
            parts = h.split(' - ', 1)
            brand = norm(parts[0].strip()) if parts else ''
            brands_in_attr.append(brand)
        q24_brand_maps[attr] = brands_in_attr

    # Read all respondent rows
    print("Reading respondent data...")
    respondents = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        respondents.append(list(row))
    total_resp = len(respondents)
    print(f"Total respondents: {total_resp}")
    wb.close()

    # === EXTRACT Q24 (all 15 attributes) ===
    # Raw data format: each Q24 attribute has 36 columns. Cell VALUES contain brand names
    # (respondent picks up to 3 brands). Count how many respondents selected each brand.
    print("\nExtracting Q24 (all 15 attributes)...")
    q24_data = {}
    for attr, start_col in Q24_ATTRS.items():
        brand_counts = {}
        base = 0
        for resp in respondents:
            # Check if respondent answered (at least one cell has a value)
            has_answer = False
            for c in range(start_col - 1, min(start_col + 35, len(resp))):
                if resp[c] is not None and str(resp[c]).strip():
                    has_answer = True
                    break
            if not has_answer:
                continue
            base += 1

            # Each cell VALUE is a brand name the respondent selected
            for c in range(start_col - 1, min(start_col + 35, len(resp))):
                val = resp[c]
                if val is not None and str(val).strip():
                    brand_name = norm(str(val).strip())
                    if is_target(brand_name):
                        brand_counts[brand_name] = brand_counts.get(brand_name, 0) + 1

        attr_data = {}
        for brand in TARGET_BRANDS:
            count = brand_counts.get(brand, 0)
            attr_data[brand] = {"Total": round(count / base, 6) if base > 0 else 0}

        q24_data[attr] = attr_data
        top = sorted(attr_data.items(), key=lambda x: -x[1]['Total'])[:3]
        top_str = ', '.join(f'{b}={v["Total"]*100:.1f}%' for b,v in top)
        print(f"  {attr}: base={base}, top3: {top_str}")

    # === EXTRACT Q15b (main brand) ===
    print("\nExtracting Q15b (main brand)...")
    main_brand_counts = {}
    main_brand_total = 0
    main_brand_per_resp = {}  # respondent index -> main brand

    for idx, resp in enumerate(respondents):
        val = resp[Q15B_COL - 1] if (Q15B_COL - 1) < len(resp) else None
        if val is not None and str(val).strip():
            brand = norm(str(val).strip())
            main_brand_total += 1
            main_brand_counts[brand] = main_brand_counts.get(brand, 0) + 1
            main_brand_per_resp[idx] = brand

    q15b_data = {}
    for brand in TARGET_BRANDS:
        count = main_brand_counts.get(brand, 0)
        q15b_data[brand] = {"count": count, "pct": round(count / main_brand_total, 6) if main_brand_total > 0 else 0}
    print(f"  Total with main brand: {main_brand_total}")
    for b, d in sorted(q15b_data.items(), key=lambda x: -x[1]['pct']):
        if d['count'] > 0:
            print(f"  {b}: {d['count']} ({d['pct']*100:.1f}%)")

    # === EXTRACT Q25 PER BRAND (cross-tab with Q15b) ===
    print("\nExtracting Q25 per brand (via Q15b cross-tab)...")
    q25_per_brand = {}
    for brand in TARGET_BRANDS:
        # Find respondents whose main brand is this brand
        brand_resps = [idx for idx, b in main_brand_per_resp.items() if b == brand]
        brand_base = len(brand_resps)
        reasons = {}
        for reason, col in Q25_COLS.items():
            count = 0
            for idx in brand_resps:
                val = respondents[idx][col - 1] if (col - 1) < len(respondents[idx]) else None
                if val is not None and str(val).strip():
                    count += 1
            reasons[reason] = {"Total": round(count / brand_base, 6) if brand_base > 0 else 0}
        q25_per_brand[brand] = {"base": brand_base, "reasons": reasons}
        loc_pct = reasons.get('Location convenience', {}).get('Total', 0)
        uniq_pct = reasons.get('Unique menu appeal', {}).get('Total', 0)
        print(f"  {brand}: base={brand_base}, Location={loc_pct*100:.1f}%, Unique={uniq_pct*100:.1f}%")

    # === EXTRACT Q3 (channel frequency) ===
    print("\nExtracting Q3 (channel frequency)...")
    q3_data = {}
    for channel, col in Q3_COLS.items():
        freq_counts = {}
        base = 0
        for resp in respondents:
            val = resp[col - 1] if (col - 1) < len(resp) else None
            if val is not None and str(val).strip():
                base += 1
                freq = str(val).strip()
                freq_counts[freq] = freq_counts.get(freq, 0) + 1
        q3_data[channel] = {
            "base": base,
            "options": {k: {"Total": round(v / base, 6)} for k, v in freq_counts.items()} if base > 0 else {}
        }
        print(f"  {channel}: base={base}, options: {list(freq_counts.keys())[:5]}")

    # === MERGE INTO EXISTING pa_data.json ===
    print(f"\nLoading existing {OUTPUT}...")
    with open(OUTPUT, 'r', encoding='utf-8') as f:
        pa_data = json.load(f)

    # Update Q24 with all 15 attributes
    pa_data['q24'] = q24_data

    # Add per-brand Q25
    pa_data['q25_per_brand'] = q25_per_brand

    # Add Q15b
    pa_data['q15b'] = q15b_data

    # Add Q3
    pa_data['q3'] = q3_data

    # Save
    with open(OUTPUT, 'w', encoding='utf-8') as f:
        json.dump(pa_data, f, indent=2, ensure_ascii=False)
    print(f"\nUpdated: {OUTPUT} ({os.path.getsize(OUTPUT)} bytes)")

    # Verification
    print("\n=== VERIFICATION ===")
    print(f"Q24 attributes: {list(pa_data['q24'].keys())}")
    print(f"Q24 Taste Kudu: {pa_data['q24']['Taste']['Kudu']['Total']*100:.1f}%")
    print(f"Q15b brands: {len(pa_data['q15b'])}")
    print(f"Q25 per brand Kudu Location: {pa_data['q25_per_brand']['Kudu']['reasons']['Location convenience']['Total']*100:.1f}%")
    print(f"Q25 per brand McDonalds Location: {pa_data['q25_per_brand']['McDonalds']['reasons']['Location convenience']['Total']*100:.1f}%")
    print(f"Q3 channels: {list(pa_data['q3'].keys())}")
    print("\nDone!")


if __name__ == "__main__":
    main()
