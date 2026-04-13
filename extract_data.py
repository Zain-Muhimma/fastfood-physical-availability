"""
Extract Physical Availability data from 'Burgerizzr data for dashboard.xlsx'
and output JSON files for the React dashboard.
"""
import json
import openpyxl
import sys
import os

sys.stdout.reconfigure(encoding='utf-8')

EXCEL_FILE = 'Burgerizzr data for dashboard.xlsx'
OUTPUT_DIR = 'public/data'

# 10 target brands (exact names as they appear in the Excel)
TARGET_BRANDS = [
    "Kudu", "Burgerizzr", "Al Baik", "McDonalds", "KFC",
    "Sign Burger", "Burger King", "Bayt Al-Shawarma", "Hardees", "Herfy"
]

# Brand name aliases (Excel uses different spellings in different places)
BRAND_ALIASES = {
    "McDonald's": "McDonalds",
    "McDonald\u2019s": "McDonalds",
    "McDonald's": "McDonalds",
    "Hardee's": "Hardees",
    "Hardee\u2019s": "Hardees",
    "Hardee's": "Hardees",
    "I'm Hungry": "Im Hungry",
    "I\u2019m Hungry": "Im Hungry",
    "Chef's": "Chefs",
    "Chef\u2019s": "Chefs",
    "Raising Cane's": "Raising Canes",
    "Raising Cane\u2019s": "Raising Canes",
    "Popeye's": "Popeyes",
    "Popeye\u2019s": "Popeyes",
    "Domino's Pizza": "Dominos Pizza",
    "Domino\u2019s Pizza": "Dominos Pizza",
    "Al Baik": "Al Baik",
    "Bayt Al-Shawarma": "Bayt Al-Shawarma",
}

def normalize_brand(name):
    """Normalize brand name to match TARGET_BRANDS."""
    name = name.strip()
    # Try direct alias first
    if name in BRAND_ALIASES:
        return BRAND_ALIASES[name]
    # Try replacing curly quotes with straight quotes
    normalized = name.replace('\u2019', "'").replace('\u2018', "'")
    if normalized in BRAND_ALIASES:
        return BRAND_ALIASES[normalized]
    # Try without any apostrophes
    no_apos = name.replace("'", "").replace("\u2019", "").replace("'", "")
    for target in TARGET_BRANDS:
        if target.replace("'", "") == no_apos:
            return target
    return name

# Demographic column indices (0-based from col C which is index 2)
# Row 3 headers: Total, Man, Woman, 16-24, 25-34, 35-45, Saudi, Non-Saudi Arab, Non-Arab,
#                Riyadh, Jeddah, Dammam/Al Khobar, Mecca/Madinah/Taif, Other, Single, Married
DEMO_COLS = {
    "Total": 2,       # Col C
    "Man": 3,          # Col D
    "Woman": 4,        # Col E
    "16-24": 5,        # Col F
    "25-34": 6,        # Col G
    "35-45": 7,        # Col H
    "Saudi": 8,        # Col I
    "Non-Saudi Arab": 9, # Col J
    "Non-Arab": 10,    # Col K
    "Riyadh": 11,      # Col L
    "Jeddah": 12,      # Col M
    "Dammam / Al Khobar": 13, # Col N
    "Mecca / Madinah / Taif": 14, # Col O
    "Other Regions": 15, # Col P
    "Single": 16,      # Col Q
    "Married": 17,     # Col R
}

def get_row_data(ws, row_num):
    """Get all values from a row as a list."""
    row = list(ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True))[0]
    return list(row)

def get_demo_values(row_data):
    """Extract demographic column values from a row."""
    result = {}
    for demo_name, col_idx in DEMO_COLS.items():
        val = row_data[col_idx] if col_idx < len(row_data) else None
        if val is not None:
            result[demo_name] = round(float(val), 6) if isinstance(val, (int, float)) else 0
        else:
            result[demo_name] = 0
    return result

def extract_brand_from_label(label):
    """Extract brand name from question label like 'q18_en__5 Kudu - ...'"""
    if not label:
        return None
    # Remove question code prefix
    parts = label.split(' ', 1)
    if len(parts) < 2:
        return None
    rest = parts[1]
    # Split on ' - ' to get brand name
    brand_parts = rest.split(' - ', 1)
    brand = brand_parts[0].strip()
    return normalize_brand(brand)

def is_target_brand(brand_name):
    """Check if brand is in our 10 target brands."""
    return normalize_brand(brand_name) in TARGET_BRANDS

def extract_per_brand_question(ws, start_row, end_row, rows_per_brand, has_mean=True):
    """
    Extract a per-brand question (Q15, Q17, Q18, Q19, Q20, Q21).
    Each brand block has: header row, optional mean row, then option rows.
    """
    result = {}
    row = start_row
    while row <= end_row:
        row_data = get_row_data(ws, row)
        label = str(row_data[0]) if row_data[0] else ""

        # Check if this is a question header row (has 'Base' in col B)
        col_b = str(row_data[1]) if row_data[1] else ""

        if col_b == "Base" and label:
            brand = extract_brand_from_label(label)
            if brand and is_target_brand(brand):
                brand_data = {"base": get_demo_values(row_data), "options": {}}

                # Read subsequent rows for this brand block
                offset = 1
                if has_mean:
                    offset = 2  # Skip mean row

                for i in range(offset, rows_per_brand):
                    opt_row = get_row_data(ws, row + i)
                    opt_label = str(opt_row[1]) if opt_row[1] else ""
                    if opt_label and opt_label not in ("Mean", "Base", ""):
                        # Clean up option label
                        opt_label = opt_label.strip()
                        brand_data["options"][opt_label] = get_demo_values(opt_row)

                result[brand] = brand_data

            row += rows_per_brand
        else:
            row += 1

    return result

def extract_multi_select_question(ws, start_row, num_rows):
    """
    Extract a multi-select question (Q16, Q24_x).
    Structure: header row (Base), then one row per brand option.
    """
    result = {}
    header = get_row_data(ws, start_row)

    for i in range(1, num_rows):
        row_data = get_row_data(ws, start_row + i)
        brand_name = str(row_data[1]) if row_data[1] else ""
        brand_name = brand_name.strip()

        if brand_name and brand_name not in ("None", ""):
            normalized = normalize_brand(brand_name)
            if is_target_brand(normalized):
                result[normalized] = get_demo_values(row_data)

    return result

def extract_q22(ws, start_row, end_row):
    """Extract Q22 (WOM) - per-brand multi-select with 5 options + None."""
    result = {}
    row = start_row
    while row <= end_row:
        row_data = get_row_data(ws, row)
        label = str(row_data[0]) if row_data[0] else ""
        col_b = str(row_data[1]) if row_data[1] else ""

        if col_b == "Base" and "q22" in label.lower():
            brand = extract_brand_from_label(label)
            if brand and is_target_brand(brand):
                brand_data = {"base": get_demo_values(row_data), "options": {}}
                # Read 5 option rows after base
                for i in range(1, 6):
                    opt_row = get_row_data(ws, row + i)
                    opt_label = str(opt_row[1]) if opt_row[1] else ""
                    if opt_label.strip():
                        brand_data["options"][opt_label.strip()] = get_demo_values(opt_row)
                result[brand] = brand_data
            row += 6
        else:
            row += 1
    return result

def extract_q25(ws, start_row):
    """Extract Q25 (main brand reasons) - single multi-select block."""
    result = {"base": {}, "reasons": {}}
    header = get_row_data(ws, start_row)
    result["base"] = get_demo_values(header)

    # Read reason rows (15 reasons)
    for i in range(1, 16):
        row_data = get_row_data(ws, start_row + i)
        reason = str(row_data[1]) if row_data[1] else ""
        reason = reason.strip()
        if reason:
            # Shorten reason labels
            short = reason.split('–')[0].strip() if '–' in reason else reason.split(' – ')[0].strip() if ' – ' in reason else reason[:50]
            result["reasons"][short] = get_demo_values(row_data)

    return result

def extract_q12(ws, start_row):
    """Extract Q12 (category choice factors)."""
    result = {"base": {}, "factors": {}}
    header = get_row_data(ws, start_row)
    result["base"] = get_demo_values(header)

    # Read factor rows (up to 18 factors)
    for i in range(1, 19):
        row_data = get_row_data(ws, start_row + i)
        if row_data[1] is None:
            break
        factor = str(row_data[1]).strip()
        if factor and factor not in ("Base", ""):
            result["factors"][factor] = get_demo_values(row_data)

    return result

def extract_q15(ws, start_row, end_row):
    """Extract Q15 (brand relationship funnel) - 5 options + mean per brand."""
    result = {}
    row = start_row
    while row <= end_row:
        row_data = get_row_data(ws, row)
        label = str(row_data[0]) if row_data[0] else ""
        col_b = str(row_data[1]) if row_data[1] else ""

        if col_b == "Base" and "q15" in label.lower():
            brand = extract_brand_from_label(label)
            if brand and is_target_brand(brand):
                brand_data = {"base": get_demo_values(row_data), "options": {}}
                # Skip mean (row+1), then read 5 funnel stages
                for i in range(2, 7):
                    opt_row = get_row_data(ws, row + i)
                    opt_label = str(opt_row[1]) if opt_row[1] else ""
                    if opt_label.strip():
                        # Shorten funnel labels
                        short = opt_label.strip()
                        if "Never heard" in short:
                            short = "Unaware"
                        elif "know the brand" in short.lower():
                            short = "Know but untried"
                        elif "tried it before" in short.lower():
                            short = "Lapsed"
                        elif "occasionally" in short.lower():
                            short = "Occasional"
                        elif "regularly" in short.lower():
                            short = "Regular"
                        brand_data["options"][short] = get_demo_values(opt_row)
                result[brand] = brand_data
            row += 7
        else:
            row += 1
    return result


def main():
    print(f"Loading {EXCEL_FILE}...")
    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)
    ws = wb['%']

    # Convert to list for random access
    print("Reading all rows...")
    all_rows = []
    for row in ws.iter_rows(values_only=True):
        all_rows.append(list(row))
    total_rows = len(all_rows)
    print(f"Total rows: {total_rows}")

    # We need to work with 1-based row access via the sheet
    # But we already have row positions from exploration

    data = {
        "meta": {
            "wave": "W3",
            "totalBase": 1201,
            "brands": TARGET_BRANDS,
            "demographics": list(DEMO_COLS.keys())
        }
    }

    # Q15 - Brand relationship funnel (rows 380-625)
    print("Extracting Q15 (Trial Penetration)...")
    data["q15"] = extract_q15(ws, 380, 625)
    print(f"  Found {len(data['q15'])} brands")

    # Q16 - Ad recall (row 662, ~37 rows)
    print("Extracting Q16 (Ad Cut-Through)...")
    data["q16"] = extract_multi_select_question(ws, 662, 37)
    print(f"  Found {len(data['q16'])} brands")

    # Q17 - Frequency momentum (rows 699-870, 5 rows per brand)
    print("Extracting Q17 (Frequency Momentum)...")
    data["q17"] = extract_per_brand_question(ws, 699, 870, 5, has_mean=True)
    print(f"  Found {len(data['q17'])} brands")

    # Q18 - Ease of dine-in (rows 874-1048, 5 rows per brand)
    print("Extracting Q18 (Ease Score / Friction Rate)...")
    data["q18"] = extract_per_brand_question(ws, 874, 1048, 5, has_mean=True)
    print(f"  Found {len(data['q18'])} brands")

    # Q19 - Overall impression (rows 1049-1223, 5 rows per brand)
    print("Extracting Q19 (Impression Score)...")
    data["q19"] = extract_per_brand_question(ws, 1049, 1223, 5, has_mean=True)
    print(f"  Found {len(data['q19'])} brands")

    # Q20 - Value for money (rows 1224-1398, 5 rows per brand)
    print("Extracting Q20 (Value Standout)...")
    data["q20"] = extract_per_brand_question(ws, 1224, 1398, 5, has_mean=True)
    print(f"  Found {len(data['q20'])} brands")

    # Q21 - Popularity trend (rows 1399-1573, 5 rows per brand)
    print("Extracting Q21 (Perceived Momentum)...")
    data["q21"] = extract_per_brand_question(ws, 1399, 1573, 5, has_mean=True)
    print(f"  Found {len(data['q21'])} brands")

    # Q22 - WOM (rows 1574-1783, 6 rows per brand)
    print("Extracting Q22 (Net Advocacy)...")
    data["q22"] = extract_q22(ws, 1574, 1783)
    print(f"  Found {len(data['q22'])} brands")

    # Q24 attributes (each is a multi-select block with ~37 brand rows)
    q24_items = {
        "Taste": 2376,
        "Price": 2450,
        "Variety": 2524,
        "Health": 2561,
        "Location": 2598,
        "Reputation": 2820,
        "Innovation": 2894,
    }
    print("Extracting Q24 (Brand Attributes)...")
    data["q24"] = {}
    for attr_name, start in q24_items.items():
        data["q24"][attr_name] = extract_multi_select_question(ws, start, 37)
        print(f"  {attr_name}: {len(data['q24'][attr_name])} brands")

    # Q25 - Main brand reasons (row 2931)
    print("Extracting Q25 (Main Brand Reasons)...")
    data["q25"] = extract_q25(ws, 2931)
    print(f"  Found {len(data['q25']['reasons'])} reasons")

    # Q12 - Category choice factors (row 353)
    print("Extracting Q12 (Category Choice Factors)...")
    data["q12"] = extract_q12(ws, 353)
    print(f"  Found {len(data['q12']['factors'])} factors")

    wb.close()

    # Output JSON
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    output_path = os.path.join(OUTPUT_DIR, 'pa_data.json')
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    print(f"\nWritten: {output_path} ({os.path.getsize(output_path)} bytes)")

    # Create brands.json
    brands = [
        {"id": "kudu", "name": "Kudu", "logo": "/logos/kudu.png", "isDefault": True},
        {"id": "burgerizzr", "name": "Burgerizzr", "logo": "/logos/burgerizzr.png", "isDefault": False},
        {"id": "albaik", "name": "Al Baik", "logo": "/logos/albaik.png", "isDefault": False},
        {"id": "mcdonalds", "name": "McDonalds", "logo": "/logos/mcdonalds.png", "isDefault": False},
        {"id": "kfc", "name": "KFC", "logo": "/logos/kfc.png", "isDefault": False},
        {"id": "signburger", "name": "Sign Burger", "logo": "/logos/signburger.png", "isDefault": False},
        {"id": "burgerking", "name": "Burger King", "logo": "/logos/burgerking.png", "isDefault": False},
        {"id": "baytalshawarma", "name": "Bayt Al-Shawarma", "logo": "/logos/baytalshawarma.png", "isDefault": False},
        {"id": "hardees", "name": "Hardees", "logo": "/logos/hardees.png", "isDefault": False},
        {"id": "herfy", "name": "Herfy", "logo": "/logos/herfy.png", "isDefault": False},
    ]
    brands_path = os.path.join(OUTPUT_DIR, 'brands.json')
    with open(brands_path, 'w', encoding='utf-8') as f:
        json.dump(brands, f, indent=2, ensure_ascii=False)
    print(f"Written: {brands_path}")

    # Verify key data points
    print("\n=== VERIFICATION ===")
    if "Kudu" in data.get("q18", {}):
        ease = data["q18"]["Kudu"]["options"]
        for opt, vals in ease.items():
            print(f"  Q18 Kudu {opt}: Total={vals.get('Total', 'N/A')}")

    if "McDonalds" in data.get("q16", {}):
        print(f"  Q16 McDonalds Ad Recall Total: {data['q16']['McDonalds'].get('Total', 'N/A')}")

    print("\nDone!")


if __name__ == "__main__":
    main()
